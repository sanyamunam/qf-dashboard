# -*- coding: utf-8 -*-
"""KPI Mapping - OBS merged workbook -> src/data/obs.json.

Sheet 'Actuals & Targets', 241 data rows. The Executive Dashboard reads ONLY
the rows where Dashboard = Executive AND the KPI Name cell carries the blue
highlight — exactly ten: 2, 5, 11, 21, 28, 48, 61, 66, 81, 85. The parser
records the highlight as a flag rather than hard-coding the ten, and asserts
the ten anyway so a silent re-colouring of the sheet cannot pass unnoticed.

Numbers stay numbers; strings like 'TBU' are preserved as notes, never
coerced and never zero. No unit lifting here — normalisation happens once,
on read, in src/model/obs.ts, where the rule can be documented next to its
consumers.
"""
import json
import math
import re
from pathlib import Path

import openpyxl

# The OneDrive copy is the updated sheet (Policy Hub Insights standalone);
# the Downloads copy still carries the older nesting. OneDrive's lock throws
# PermissionError on in-place reads, so the file is copied to temp first.
import shutil
import subprocess
import tempfile

# 26 Aug 2026: the client re-issued the workbook as '- Updated'. It is the
# one the chart-system brief was written against - verified independently on
# five counts the brief cites (targets 30/43/42/47 by year, 22 rows carrying
# both actual and target across 2022-25, 174 with all three future targets,
# 9 green-polarity zero-target rows, 15 Polarity: Red). The previous file
# matched none of them.
ONEDRIVE = Path(r"C:\Users\user\OneDrive - applab.qa\QF Observatory\KPI Mapping - OBS merged - Updated.xlsx")
# a project-local snapshot of the SAME updated sheet, so a locked or
# dehydrated OneDrive file never silently changes what gets parsed
SNAPSHOT = Path(__file__).resolve().parent.parent / "data" / "KPI Mapping - OBS merged (snapshot 2026-08-26 updated).xlsx"


def _resolve_source() -> Path:
    """The updated workbook, or a faithful copy of it - never anything else.

    This used to fall back to an OLDER workbook when OneDrive refused a read,
    and that is precisely how a stale parse reached the build: the variance
    rows came through empty and nobody was told. Excel holds a lock that
    `shutil` cannot pass but PowerShell can, so that is tried next; if every
    route fails the parser STOPS rather than parse the wrong sheet.
    """
    tmp = Path(tempfile.gettempdir()) / "obs-kpi-mapping.xlsx"
    try:
        shutil.copyfile(ONEDRIVE, tmp)
        return tmp
    except OSError:
        pass
    try:
        subprocess.run(
            ["powershell", "-NoProfile", "-Command", f'Copy-Item -LiteralPath "{ONEDRIVE}" -Destination "{tmp}" -Force'],
            check=True,
            capture_output=True,
        )
        print("note: OneDrive file is locked - copied via PowerShell")
        return tmp
    except Exception:
        pass
    if SNAPSHOT.exists():
        print(f"note: live workbook unreadable - using {SNAPSHOT.name}")
        return SNAPSHOT
    raise SystemExit(
        "STOP: cannot read the updated workbook and no snapshot exists.\n"
        "      Refusing to parse an older sheet - close Excel and re-run."
    )


SRC = _resolve_source()
OUT = Path(__file__).resolve().parent.parent / "src" / "data" / "obs.json"

YEARS_A = ["2022", "2023", "2024", "2025"]
YEARS_T = ["2022", "2023", "2024", "2025", "2026", "2027", "2028"]


def num(v):
    """A number, whether the cell stored it as one or as text.

    The updated workbook types whole swathes of the variance and rate rows as
    TEXT - '-28', '50', '10', '8.45'. Reading only real numerics silently
    emptied every one of them, which is why the variance KPIs came through with
    no history at all. Anything that parses as a number IS one; 'TBU' and 'NR'
    still are not, and fall through to `note` as unset.
    """
    if isinstance(v, (int, float)) and not isinstance(v, bool) and not (isinstance(v, float) and math.isnan(v)):
        return v
    if isinstance(v, str):
        t = v.strip().replace(",", "").replace("%", "").replace("\u2212", "-")
        if re.fullmatch(r"-?\d+(?:\.\d+)?", t):
            return float(t)
    return None


def note(v):
    if v is None or num(v) is not None:
        return None
    s = str(v).strip()
    return s or None


def txt(v):
    if v is None:
        return None
    s = str(v).replace("\xa0", " ").strip()
    return s or None


wb = openpyxl.load_workbook(SRC)  # styles needed for the highlight flag
wsx = wb["Actuals & Targets"]
wbv = openpyxl.load_workbook(SRC, data_only=True)
ws = wbv["Actuals & Targets"]

rows = []
for r in range(2, ws.max_row + 1):
    name = txt(ws.cell(r, 7).value)
    if not name:
        continue

    def c(col):
        return ws.cell(r, col).value

    category = txt(c(4)) or ""
    # 'Group - Subgroup'; a category without the dash is its own group
    group, _, sub = (p.strip() for p in category.partition(" - "))
    # the fill sits on the Dashboard AND KPI Name columns; either counts
    highlighted = wsx.cell(r, 1).fill.patternType == "solid" or wsx.cell(r, 7).fill.patternType == "solid"

    rows.append(
        {
            "row": r,
            "dashboard": txt(c(1)),
            "entity": txt(c(2)),
            # the cleaner of the two entity columns — what the card shows
            "proposedEntity": txt(c(3)),
            "category": category,
            "group": group,
            "subgroup": sub or group,
            "theme": txt(c(5)),
            "framework": txt(c(6)),
            "name": name,
            "definition": txt(c(8)),
            "polarity": txt(c(9)),
            "highlighted": highlighted,
            "actuals": {y: num(c(10 + i)) for i, y in enumerate(YEARS_A)},
            "actualNotes": {y: note(c(10 + i)) for i, y in enumerate(YEARS_A)},
            "q1": num(c(14)),
            "q1Note": note(c(14)),
            "targets": {y: num(c(15 + i)) for i, y in enumerate(YEARS_T)},
            "targetNotes": {y: note(c(15 + i)) for i, y in enumerate(YEARS_T)},
        }
    )

by = {x["row"]: x for x in rows}

# ── the ten, pinned by name so a re-sort or re-colour cannot slip through ──
DASH = [r["row"] for r in rows if (r["dashboard"] or "").lower().startswith("exec") and r["highlighted"]]
assert DASH == [2, 5, 11, 21, 28, 48, 61, 66, 81, 85], DASH
assert by[2]["name"] == "Budget Variance" and by[2]["q1"] == 18 and by[2]["actuals"]["2025"] == 0.1
assert by[5]["name"] == "Qatarization" and by[5]["targets"]["2024"] == 0.4 and by[5]["targets"]["2025"] == 0.25
assert by[11]["name"] == "Total Policy Adoptions" and by[11]["q1"] == 2 and by[11]["targets"]["2026"] == 6
assert by[11]["entity"] is None and by[11]["theme"] is None and by[11]["framework"] is None
# the UPDATED sheet: Policy Hub Insights is standalone, not nested
assert by[11]["category"] == "Policy Hub Insights", by[11]["category"]
assert by[40]["name"].startswith("Patents Granted") and by[40]["category"] == ""
assert by[21]["name"] == "Footfall - EC total" and by[21]["q1"] == 486119
assert by[28]["name"] == "Active QPHI projects" and by[28]["actuals"]["2025"] == 250 and by[28]["q1"] == 28
assert by[48]["actuals"]["2025"] == 1027 and by[61]["actuals"]["2025"] == 612
assert by[66]["actuals"]["2025"] == 4017 and by[81]["actuals"]["2025"] == 8615
assert by[85]["actuals"]["2025"] == 0.72 and by[85]["targets"]["2026"] == 0.8

# the three rows whose Polarity column is inverted — flagged, not fixed
# (they are outside the ten; the column is authoritative for the dashboard)
assert by[4]["polarity"] == "Red" and "Revenue" in by[4]["name"]
assert by[8]["polarity"] == "Green" and by[8]["name"] == "Employee Turnover"
assert by[88]["polarity"] == "Green" and "Not Employed" in by[88]["name"]

OUT.write_text(json.dumps({"rows": rows}, indent=1), encoding="utf-8")
print(f"wrote {len(rows)} rows -> {OUT.name} · dashboard ten: {DASH}")
print("inverted-polarity rows flagged: 4 (Revenue Generated), 8 (Employee Turnover), 88 (Graduates Not Employed)")
