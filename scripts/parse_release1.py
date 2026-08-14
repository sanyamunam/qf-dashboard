"""Release 1 (executive) workbook -> src/data/exec.json.

Sheet 'Q1_updates_drill downKPIs': header row 2, data rows 3-71 (69 KPIs).
Columns: A framework, B theme, C entity, D level, E category, F name,
G definition, H source, I frequency, J-L Jan/Feb/Mar, M 2026 Q1,
N-Q actuals 2022-2025, R-X targets 2022-2028, Y/Z missing flags, AA comment.

Numbers stay numbers; strings like 'NR', 'TBU', 'Reported annually' are
preserved as notes, never coerced. Unit-suspect rows are the brief's problem
to name, not the parser's to hide.
"""
import json
import math
from pathlib import Path

import openpyxl

SRC = Path(r"C:\Users\user\OneDrive - applab.qa\QF Observatory\Al Mishkat - Release 1 KPIs.xlsx")
OUT = Path(__file__).resolve().parent.parent / "src" / "data" / "exec.json"


def num(v):
    if isinstance(v, (int, float)) and not isinstance(v, bool) and not (isinstance(v, float) and math.isnan(v)):
        return v
    return None


def note(v):
    if v is None or num(v) is not None:
        return None
    s = str(v).strip()
    return s or None


def txt(v):
    if v is None:
        return None
    s = str(v).replace("\xa0", " ").strip()
    return s or None


wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb["Q1_updates_drill downKPIs"]

rows, warnings = [], []
for r in range(3, 72):
    name = txt(ws.cell(r, 6).value)
    if not name:
        warnings.append(f"row {r}: empty name, skipped")
        continue

    def c(col):
        return ws.cell(r, col).value

    rows.append(
        {
            "row": r,
            "framework": txt(c(1)),
            "theme": txt(c(2)),
            "entity": txt(c(3)),
            "level": txt(c(4)),
            "category": txt(c(5)),
            "name": name,
            "definition": txt(c(7)),
            "source": txt(c(8)),
            "frequency": txt(c(9)),
            "monthly": {"jan": num(c(10)), "feb": num(c(11)), "mar": num(c(12))},
            "q1": num(c(13)),
            "q1Note": note(c(13)),
            "actuals": {str(2022 + i): num(c(14 + i)) for i in range(4)},
            "actualNotes": {str(2022 + i): note(c(14 + i)) for i in range(4)},
            "targets": {str(2022 + i): num(c(18 + i)) for i in range(7)},
            "targetNotes": {str(2022 + i): note(c(18 + i)) for i in range(7)},
            "comment": txt(c(27)),
        }
    )

# ---- hard assertions: every figure the brief will print, against its cell ----
by = {x["row"]: x for x in rows}
assert len(rows) == 69, len(rows)

f = by[56]  # EC footfall, monthly
assert f["monthly"] == {"jan": 211772, "feb": 237801, "mar": 36546}, f["monthly"]
assert f["q1"] == 486119 and f["q1"] == sum(f["monthly"].values())
assert f["actuals"]["2025"] == 3051433

v = by[69]  # vacancies, monthly
assert v["monthly"] == {"jan": 610, "feb": 673, "mar": 695}, v["monthly"]
assert max(x for x in v["actuals"].values() if x is not None) == 620  # 695 beats every year-end

lv = by[66]  # leadership vacancies
assert lv["monthly"] == {"jan": 15, "feb": 14, "mar": 14} and lv["actuals"]["2025"] == 19

rev = by[65]  # revenue generated, QAR millions, cumulative
assert rev["monthly"] == {"jan": 171, "feb": 342, "mar": 545}, rev["monthly"]
assert rev["q1"] == 545 and rev["actuals"]["2025"] == 2055

g = by[3]  # graduates employed
assert [g["actuals"][y] for y in ("2022", "2023", "2024", "2025")] == [0.54, 0.65, 0.65, 0.72]
assert g["targets"]["2026"] == 0.8

assert by[49]["q1"] == 38683 and by[49]["targets"]["2026"] == 30000  # genomes (cumulative caveat)
assert by[62]["q1"] == 0 and by[62]["actuals"]["2025"] == 169546  # cultural footfall

OUT.write_text(
    json.dumps(
        {
            "generatedFrom": SRC.name,
            "sheet": "Q1_updates_drill downKPIs",
            "period": "2026Q1",
            "parseWarnings": warnings,
            "kpis": rows,
        },
        indent=1,
        ensure_ascii=False,
    ),
    encoding="utf-8",
)
print(f"wrote {len(rows)} exec KPIs -> {OUT}")
