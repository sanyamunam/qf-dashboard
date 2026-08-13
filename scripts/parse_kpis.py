# -*- coding: utf-8 -*-
"""Parse Al Mishkat Release 2 KPIs.xlsx (ALL sheet) into the typed JSON model.

Defensive against: 'NR' strings, '700+', comma-formatted numbers, stray 'x'/backtick,
trailing blank rows, blank chart-type cells inheriting the previous Name: group.
Emits src/data/kpis.json + a verification report to stdout.
"""
import json, re, sys, io
from pathlib import Path
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "data" / "release2-kpis.xlsx"
OUT = ROOT / "src" / "data" / "kpis.json"

ACTUAL_YEARS = ["2022", "2023", "2024", "2025", "2026Q1"]
TARGET_YEARS = ["2022", "2023", "2024", "2025", "2026", "2027", "2028"]

parse_warnings = []

def clean_num(v, ctx=""):
    """Return (value:float|None, raw:str|None, flag:str|None). Never coerce NR to 0."""
    if v is None:
        return None, None, None
    if isinstance(v, (int, float)):
        return float(v), None, None
    s = str(v).strip()
    if s == "" or s in ("`", "x", "-"):
        if s:
            parse_warnings.append(f"stray value {s!r} at {ctx}")
        return None, s if s else None, "stray" if s else None
    if s.upper() == "NR":
        return None, "NR", "not_reported"
    if s.upper() in ("NA", "N/A"):
        return None, "NA", "not_available"
    if s.endswith("+"):
        try:
            return float(s[:-1].replace(",", "")), s, "approx_plus"
        except ValueError:
            pass
    try:
        return float(s.replace(",", "")), None, None
    except ValueError:
        parse_warnings.append(f"unparseable {s!r} at {ctx}")
        return None, s, "unparseable"

def parse_chart_type(raw):
    """Split 'bar chart\\nName: Partnerships' -> (chart, group)."""
    if raw is None:
        return None, None
    s = str(raw).strip()
    m = re.search(r"name\s*:\s*(.+)", s, re.IGNORECASE)
    group = m.group(1).strip() if m else None
    chart = re.split(r"\n|name\s*:", s, flags=re.IGNORECASE)[0].strip().rstrip("\\n ").strip()
    return (chart or None), group

wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb["ALL"]
headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
idx = {h: i for i, h in enumerate(headers)}

def col(row, name):
    i = idx.get(name)
    return row[i] if i is not None else None

kpis = []
last_group = None
last_chart = None
for r, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    name = col(row, "KPI Name")
    entity = col(row, "Entity Name")
    if not name and not entity:
        continue  # trailing blank rows
    l1_raw = col(row, "L1 chart Type")
    l1_chart, group = parse_chart_type(l1_raw)
    if l1_chart:
        last_chart, last_group = l1_chart, group
    else:
        # blank chart cell inherits the previous named group
        l1_chart, group = last_chart, last_group
    actuals, targets = {}, {}
    actual_cols = ["Actual 2022", "Actual 2023", "Actual 2024", "Actual 2025", "Actual  2026 Q1"]
    for y, c in zip(ACTUAL_YEARS, actual_cols):
        val, raw, flag = clean_num(col(row, c), f"{name} {c}")
        actuals[y] = {"value": val, "raw": raw, "flag": flag}
    for y in TARGET_YEARS:
        val, raw, flag = clean_num(col(row, f"target {y}"), f"{name} target {y}")
        targets[y] = {"value": val, "raw": raw, "flag": flag}

    polarity = str(col(row, "Polarity") or "Green").strip()
    reporting = str(col(row, "Reporting Period") or "").strip() or None
    annual = bool(reporting and "annual" in reporting.lower())

    # cadence: continuous | annual | biennial | summit-year
    tvals = [targets[y]["value"] for y in ["2026", "2027", "2028"]]
    known = [t for t in tvals if t is not None]
    cadence = "continuous"
    if annual:
        cadence = "annual"
    if len(known) >= 2 and any(t == 0 for t in known) and any(t and t > 0 for t in known):
        cadence = "cyclical"  # biennial / summit-year — off-years have 0 targets
    kpi = {
        "id": f"kpi-{r-1}",
        "row": r,
        "entity": str(entity).strip() if entity else None,
        "polarity": polarity,
        "category": (str(col(row, "KPI Category")).strip() if col(row, "KPI Category") else None),
        "theme": (str(col(row, "Thematic Area")).strip() if col(row, "Thematic Area") else None),
        "framework": (str(col(row, "Performance Framework")).strip() if col(row, "Performance Framework") else None),
        "name": str(name).strip() if name else None,
        "definition": (str(col(row, "KPI Definition")).strip() if col(row, "KPI Definition") else None),
        "l1Chart": l1_chart,
        "chartGroup": group,
        "l2Chart": (str(col(row, "L2 chart Type")).strip() if col(row, "L2 chart Type") else None),
        "actuals": actuals,
        "targets": targets,
        "reportingPeriod": reporting,
        "cadence": cadence,
    }

    # derived
    a26 = actuals["2026Q1"]["value"]
    t26 = targets["2026"]["value"]
    nr26 = actuals["2026Q1"]["raw"] == "NR"
    hist = [(y, actuals[y]["value"]) for y in ACTUAL_YEARS if actuals[y]["value"] is not None]
    kpi["historyPoints"] = len(hist)
    kpi["hasEnoughHistoryForLine"] = len(hist) >= 3

    # six-state engine (§5.2.1)
    if t26 is None:
        state = "NO_TARGET_SET"
    elif t26 == 0:
        state = "IDLE_THIS_CYCLE"
    elif a26 is None or nr26:
        state = "REPORTS_AT_YEAR_END" if annual else "NOT_REPORTED"
    elif annual:
        state = "REPORTS_AT_YEAR_END"
    elif polarity == "Red":
        state = "ABOVE_CEILING" if a26 > t26 else "WITHIN_LIMIT"
    elif a26 >= t26:
        state = "TARGET_ALREADY_MET"
    else:
        state = "IN_PROGRESS"
    kpi["state"] = state

    # largest proportional change across trustworthy readings (selection rule #1).
    # Annual/cyclical KPIs carry a literal 0 in the Q1 cell that is a not-yet-reported
    # artifact, not a collapse — movement uses annual history only for those; cyclical
    # KPIs additionally compare non-zero (on-cycle) readings.
    if cadence == "continuous":
        mv = hist
    else:
        mv = [(y, v) for y, v in hist if y != "2026Q1"]
    # a trailing Q1 zero is "nothing reported yet this quarter", never proof of decline
    if mv and mv[-1][0] == "2026Q1" and mv[-1][1] == 0:
        mv = mv[:-1]
    if cadence == "cyclical":
        mv = [(y, v) for y, v in mv if v != 0]
    vals = [v for _, v in mv]
    kpi["movementSeries"] = mv
    kpi["propChange"] = None
    kpi["movementScore"] = None
    if len(vals) >= 2:
        base = max(abs(v) for v in vals)
        if base > 0:
            import math
            kpi["propChange"] = (vals[-1] - vals[0]) / base
            # weight by real-world scale so 10,000 people outrank a 5-point % flip
            kpi["movementScore"] = abs(kpi["propChange"]) * math.log10(1 + base)
    # exact-hit + overshoot flags
    kpi["exactHit"] = (a26 is not None and t26 not in (None, 0) and a26 == t26)
    kpi["overshoot"] = (a26 is not None and t26 not in (None, 0) and polarity == "Green" and a26 > 2 * t26)

    kpis.append(kpi)

# ---------- verification report ----------
from collections import Counter, defaultdict
print(f"rows: {len(kpis)}  (brief says 151)")
print("themes:", dict(Counter(k['theme'] for k in kpis)))
print("framework:", dict(Counter(k['framework'] for k in kpis)))
print("polarity:", dict(Counter(k['polarity'] for k in kpis)))
print("entities:", dict(Counter(k['entity'] for k in kpis)))
print("states:", dict(Counter(k['state'] for k in kpis)))
st = defaultdict(Counter)
for k in kpis:
    st[k['theme']][k['state']] += 1
for t, c in st.items():
    print(f"  {t}: {dict(c)}")
print("annually reported:", sum(1 for k in kpis if k['cadence'] in ('annual',)),
      " cyclical:", sum(1 for k in kpis if k['cadence'] == 'cyclical'))
groups = Counter(k['chartGroup'] for k in kpis if k['chartGroup'])
print(f"chart groups: {len(groups)} distinct; multi-KPI groups: {[(g, n) for g, n in groups.items() if n > 1]}")
print("zero Q1 actuals:", sum(1 for k in kpis if k['actuals']['2026Q1']['value'] == 0))
print("exact hits:", sum(1 for k in kpis if k['exactHit']), " overshoots>2x:", sum(1 for k in kpis if k['overshoot']))
wish = [k for k in kpis if k['name'] and 'eneficiar' in k['name'] and k['entity'] == 'WISH']
for k in wish:
    print("WISH beneficiaries:", [(y, k['actuals'][y]['value']) for y in ACTUAL_YEARS], "t26:", k['targets']['2026']['value'])
top = sorted([k for k in kpis if k.get('movementScore') is not None], key=lambda k: -k['movementScore'])[:10]
print("largest movements (scale-weighted):")
for k in top:
    print(f"  {k['entity']} | {k['name'][:50]} | {k['propChange']:+.1%} score={k['movementScore']:.2f} | {[(y, v) for y, v in k['movementSeries']]}")
print("parse warnings:", len(parse_warnings))
for w in parse_warnings[:15]:
    print("  !", w)

OUT.parent.mkdir(parents=True, exist_ok=True)
model = {
    "generatedFrom": "Al Mishkat - Release 2 KPIs (1).xlsx / ALL sheet",
    "period": "2026-Q1",
    "parseWarnings": parse_warnings,
    "kpis": kpis,
}
OUT.write_text(json.dumps(model, indent=1), encoding="utf-8")
print("wrote", OUT, f"{OUT.stat().st_size/1024:.0f}KB")
